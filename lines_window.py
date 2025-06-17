from PyQt5.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QLineEdit, 
    QTableWidget, QTableWidgetItem, QMessageBox, QDialog,
    QToolBar, QAction, QFileDialog, QMenu, QApplication,
    QSizePolicy, QStyle, QFormLayout, QPushButton, QSpacerItem,
    QComboBox
)
from PyQt5.QtCore import Qt, QPoint, QTimer, QTimer
from PyQt5.QtGui import QFont, QIcon
from database import Database
import pandas as pd
import os
import logging
import traceback
import openpyxl
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side

# Constants
FONT_FAMILY = "Vazir"
FONT_SIZE_NORMAL = 12
FONT_SIZE_LARGE = 18
FONT_SIZE_SMALL = 10
PADDING = 5
MARGIN = 20

# تنظیم لاگ برای دیباگ
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

# -----------------------------------------------
# Filter Popover: A small widget for column-specific filtering
# -----------------------------------------------

class FilterPopover(QWidget):
    def __init__(self, parent, column_index, column_name, current_filter, on_filter_changed):
        super().__init__(parent)
        self.setWindowFlags(Qt.Popup)
        self.column_index = column_index
        self.column_name = column_name
        self.on_filter_changed = on_filter_changed
        self.setLayout(QVBoxLayout())
        self.layout().setContentsMargins(5, 5, 5, 5)
        self.setStyleSheet("""
            QWidget {
                background-color: white; 
                border: 1px solid #ccc; 
                border-radius: 4px;
            }
            QLineEdit {
                padding: 5px;
                border: 1px solid #aaa;
                border-radius: 4px;
                min-height: 25px;
            }
            QLineEdit:focus {
                border-color: #0078d4;
            }
        """)
        
        # Initialize and set up the debounce timer
        self.filter_timer = QTimer(self)
        self.filter_timer.setSingleShot(True)
        self.filter_timer.setInterval(300)  # 300ms debounce delay
        self.filter_timer.timeout.connect(self.apply_filter_delayed)

        # Filter input with improved styling and placeholder
        self.filter_input = QLineEdit(self)
        self.filter_input.setFont(QFont("Vazir", 10))
        self.filter_input.setPlaceholderText(f"فیلتر {column_name}")
        self.filter_input.setFixedWidth(200)
        self.filter_input.setText(current_filter or "")
        self.filter_input.textChanged.connect(self.handle_text_changed)
        
        # Add input to layout
        self.layout().addWidget(self.filter_input)
        self.filter_input.setFocus()  # Auto-focus the input
        
        # Adjust size to content
        self.adjustSize()
    
    def handle_text_changed(self):
        try:
            # Reset and start the debounce timer
            self.filter_timer.stop()
            self.filter_timer.start()
        except Exception as e:
            logging.error(f"Error in handle_text_changed: {str(e)}\n{traceback.format_exc()}")
            self.show_error("خطا در پردازش فیلتر")

    def apply_filter_delayed(self):
        try:
            filter_text = self.filter_input.text().strip()
            logging.debug(f"Applying filter for column {self.column_name} ({self.column_index}): {filter_text}")
            self.on_filter_changed(self.column_index, filter_text)
        except Exception as e:
            logging.error(f"Error in apply_filter_delayed: {str(e)}\n{traceback.format_exc()}")
            self.show_error("خطا در اعمال فیلتر")

    def show_error(self, message):
        QMessageBox.critical(self, "خطا", message)

    def focusOutEvent(self, event):
        try:
            # Ensure any pending filter is applied before closing
            if self.filter_timer.isActive():
                self.filter_timer.stop()
                self.apply_filter_delayed()
            super().focusOutEvent(event)
            self.close()
        except Exception as e:
            logging.error(f"Error in focusOutEvent: {str(e)}\n{traceback.format_exc()}")

    def closeEvent(self, event):
        try:
            # Clean up timer on close
            if self.filter_timer.isActive():
                self.filter_timer.stop()
            super().closeEvent(event)
        except Exception as e:
            logging.error(f"Error in closeEvent: {str(e)}\n{traceback.format_exc()}")

# -----------------------------------------------
# Input Dialog: A dialog for adding or editing line information
# -----------------------------------------------

class LineInputDialog(QDialog):
    def __init__(self, parent=None, line_data=None, is_edit=False):
        super().__init__(parent)
        self.setWindowTitle("افزودن خط" if not is_edit else "ویرایش خط")
        self.setFixedSize(400, 800)
        self.setLayoutDirection(Qt.RightToLeft)
        self.is_edit = is_edit
        self.current_id = line_data.get('id') if line_data else None
        self.db = Database()

        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 10, 10, 10)
        self.layout.setSpacing(5)

        self.form_layout = QFormLayout()
        self.form_layout.setLabelAlignment(Qt.AlignRight)
        self.form_layout.setFormAlignment(Qt.AlignRight)
        self.form_layout.setSpacing(4)

        font = QFont("Arial", 14)

        self.line_code = QLineEdit()
        self.line_code.setFont(font)
        self.line_code.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        self.line_code.setMaxLength(20)
        line_code_layout = QHBoxLayout()
        line_code_layout.addWidget(self.line_code)
        line_code_layout.addStretch()
        self.form_layout.addRow(QLabel("کد خط:"), line_code_layout)

        self.voltage_level = QLineEdit()
        self.voltage_level.setFont(font)
        self.voltage_level.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        self.voltage_level.setMaxLength(3)
        voltage_layout = QHBoxLayout()
        voltage_layout.addWidget(self.voltage_level)
        voltage_layout.addStretch()
        self.form_layout.addRow(QLabel("سطح ولتاژ:"), voltage_layout)

        self.line_name = QLineEdit()
        self.line_name.setFont(font)
        self.line_name.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        self.line_name.setMaxLength(50)
        line_name_layout = QHBoxLayout()
        line_name_layout.addWidget(self.line_name)
        line_name_layout.addStretch()
        self.form_layout.addRow(QLabel("نام خط:"), line_name_layout)

        self.dispatch_code = QLineEdit()
        self.dispatch_code.setFont(font)
        self.dispatch_code.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        dispatch_layout = QHBoxLayout()
        dispatch_layout.addWidget(self.dispatch_code)
        dispatch_layout.addStretch()
        self.form_layout.addRow(QLabel("کد دیسپاچینگ:"), dispatch_layout)

        self.total_towers = QLineEdit()
        self.total_towers.setFont(font)
        self.total_towers.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        total_towers_layout = QHBoxLayout()
        total_towers_layout.addWidget(self.total_towers)
        total_towers_layout.addStretch()
        self.form_layout.addRow(QLabel("تعداد کل دکل‌ها:"), total_towers_layout)

        self.tension_towers = QLineEdit()
        self.tension_towers.setFont(font)
        self.tension_towers.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        tension_towers_layout = QHBoxLayout()
        tension_towers_layout.addWidget(self.tension_towers)
        tension_towers_layout.addStretch()
        self.form_layout.addRow(QLabel("دکل‌های کششی:"), tension_towers_layout)

        self.suspension_towers = QLineEdit()
        self.suspension_towers.setFont(font)
        self.suspension_towers.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        suspension_towers_layout = QHBoxLayout()
        suspension_towers_layout.addWidget(self.suspension_towers)
        suspension_towers_layout.addStretch()
        self.form_layout.addRow(QLabel("دکل‌های آویزه:"), suspension_towers_layout)

        self.line_length = QLineEdit()
        self.line_length.setFont(font)
        self.line_length.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        line_length_layout = QHBoxLayout()
        line_length_layout.addWidget(self.line_length)
        line_length_layout.addStretch()
        self.form_layout.addRow(QLabel("طول خط:"), line_length_layout)

        self.circuit_length = QLineEdit()
        self.circuit_length.setFont(font)
        self.circuit_length.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        circuit_length_layout = QHBoxLayout()
        circuit_length_layout.addWidget(self.circuit_length)
        circuit_length_layout.addStretch()
        self.form_layout.addRow(QLabel("طول مدار:"), circuit_length_layout)

        self.plain_area = QLineEdit()
        self.plain_area.setFont(font)
        self.plain_area.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        plain_area_layout = QHBoxLayout()
        plain_area_layout.addWidget(self.plain_area)
        plain_area_layout.addStretch()
        self.form_layout.addRow(QLabel("دشت:"), plain_area_layout)

        self.semi_mountainous = QLineEdit()
        self.semi_mountainous.setFont(font)
        self.semi_mountainous.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        semi_mountainous_layout = QHBoxLayout()
        semi_mountainous_layout.addWidget(self.semi_mountainous)
        semi_mountainous_layout.addStretch()
        self.form_layout.addRow(QLabel("نیمه کوهستانی:"), semi_mountainous_layout)

        self.rough_terrain = QLineEdit()
        self.rough_terrain.setFont(font)
        self.rough_terrain.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        rough_terrain_layout = QHBoxLayout()
        rough_terrain_layout.addWidget(self.rough_terrain)
        rough_terrain_layout.addStretch()
        self.form_layout.addRow(QLabel("صعب العبور:"), rough_terrain_layout)

        self.supervisor = QLineEdit()
        self.supervisor.setFont(font)
        self.supervisor.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        supervisor_layout = QHBoxLayout()
        supervisor_layout.addWidget(self.supervisor)
        supervisor_layout.addStretch()
        self.form_layout.addRow(QLabel("ناظر:"), supervisor_layout)

        self.team_leader = QComboBox()
        self.team_leader.setFont(font)
        self.team_leader.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        team_leader_layout = QHBoxLayout()
        team_leader_layout.addWidget(self.team_leader)
        team_leader_layout.addStretch()
        self.form_layout.addRow(QLabel("سرپرست اکیپ:"), team_leader_layout)
        self.load_team_leaders()

        self.operation_year = QLineEdit()
        self.operation_year.setFont(font)
        self.operation_year.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        operation_year_layout = QHBoxLayout()
        operation_year_layout.addWidget(self.operation_year)
        operation_year_layout.addStretch()
        self.form_layout.addRow(QLabel("سال بهره‌برداری:"), operation_year_layout)

        self.wire_type = QLineEdit()
        self.wire_type.setFont(font)
        self.wire_type.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        wire_type_layout = QHBoxLayout()
        wire_type_layout.addWidget(self.wire_type)
        wire_type_layout.addStretch()
        self.form_layout.addRow(QLabel("نوع سیم:"), wire_type_layout)

        self.tower_type = QLineEdit()
        self.tower_type.setFont(font)
        self.tower_type.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        tower_type_layout = QHBoxLayout()
        tower_type_layout.addWidget(self.tower_type)
        tower_type_layout.addStretch()
        self.form_layout.addRow(QLabel("نوع دکل:"), tower_type_layout)

        self.bundle_count = QLineEdit()
        self.bundle_count.setFont(font)
        self.bundle_count.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        bundle_count_layout = QHBoxLayout()
        bundle_count_layout.addWidget(self.bundle_count)
        bundle_count_layout.addStretch()
        self.form_layout.addRow(QLabel("تعداد باندل:"), bundle_count_layout)

        self.circuit_count = QLineEdit()
        self.circuit_count.setFont(font)
        self.circuit_count.setStyleSheet("padding: 5px; border: 1px solid #ccc; background-color: white; width: 220px; border-radius: 6px;")
        circuit_count_layout = QHBoxLayout()
        circuit_count_layout.addWidget(self.circuit_count)
        circuit_count_layout.addStretch()
        self.form_layout.addRow(QLabel("تعداد مدار:"), circuit_count_layout)

        if line_data:
            self.line_code.setText(str(line_data.get('line_code', '')))
            self.voltage_level.setText(str(line_data.get('voltage_level', '')))
            self.line_name.setText(str(line_data.get('line_name', '')))
            self.dispatch_code.setText(str(line_data.get('dispatch_code', '')))
            self.total_towers.setText(str(line_data.get('total_towers', '')))
            self.tension_towers.setText(str(line_data.get('tension_towers', '')))
            self.suspension_towers.setText(str(line_data.get('suspension_towers', '')))
            self.line_length.setText(str(line_data.get('line_length', '')))
            self.circuit_length.setText(str(line_data.get('circuit_length', '')))
            self.plain_area.setText(str(line_data.get('plain_area', '')))
            self.semi_mountainous.setText(str(line_data.get('semi_mountainous', '')))
            self.rough_terrain.setText(str(line_data.get('rough_terrain', '')))
            self.supervisor.setText(str(line_data.get('supervisor', '')))
            current_leader = str(line_data.get('team_leader', ''))
            index = self.team_leader.findText(current_leader)
            if index >= 0:
                self.team_leader.setCurrentIndex(index)
            else:
                self.team_leader.setCurrentIndex(0)
            self.operation_year.setText(str(line_data.get('operation_year', '')))
            self.wire_type.setText(str(line_data.get('wire_type', '')))
            self.tower_type.setText(str(line_data.get('tower_type', '')))
            self.bundle_count.setText(str(line_data.get('bundle_count', '')))
            self.circuit_count.setText(str(line_data.get('circuit_count', '')))

        self.layout.addLayout(self.form_layout)
        self.layout.addSpacerItem(QSpacerItem(0, 15, QSizePolicy.Minimum, QSizePolicy.Fixed))

        self.button_layout = QHBoxLayout()
        self.save_button = QPushButton("ذخیره")
        self.save_button.setFont(font)
        self.save_button.setStyleSheet("padding: 8px 5px; background-color: #4CAF50; color: white; border: none; border-radius: 5px;")
        self.cancel_button = QPushButton("لغو")
        self.cancel_button.setFont(font)
        self.cancel_button.setStyleSheet("padding: 8px 5px; background-color: #f44336; color: white; border: none; border-radius: 5px;")
        self.button_layout.addWidget(self.save_button)
        self.button_layout.addWidget(self.cancel_button)
        self.layout.addLayout(self.button_layout)

        self.save_button.clicked.connect(self.save_line)
        self.cancel_button.clicked.connect(self.reject)

    def load_team_leaders(self):
        try:
            query = """
                SELECT first_name || ' ' || last_name
                FROM teams
                WHERE position = 'سرپرست اکیپ'
            """
            rows = self.db.fetch_all(query)
            self.team_leader.addItem("")
            for row in rows:
                self.team_leader.addItem(row[0])
        except Exception as e:
            logging.error(f"Error in load_team_leaders: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری سرپرست‌ها: {str(e)}")

    def validate_inputs(self):
        if not self.line_code.text():
            return False, "کد خط اجباری است!"
        if not self.line_name.text():
            return False, "نام خط اجباری است!"
        if not self.team_leader.currentText():
            return False, "سرپرست اکیپ اجباری است!"
        if self.total_towers.text() and not self.total_towers.text().isdigit():
            return False, "تعداد کل دکل‌ها باید عدد باشد!"
        if self.tension_towers.text() and not self.tension_towers.text().isdigit():
            return False, "تعداد دکل‌های کششی باید عدد باشد!"
        if self.suspension_towers.text() and not self.suspension_towers.text().isdigit():
            return False, "تعداد دکل‌های آویزه باید عدد باشد!"
        if self.line_length.text() and not re.match(r"^\d*\.?\d*$", self.line_length.text()):
            return False, "طول خط باید عدد باشد!"
        if self.circuit_length.text() and not re.match(r"^\d*\.?\d*$", self.circuit_length.text()):
            return False, "طول مدار باید عدد باشد!"
        if self.voltage_level.text() and not self.voltage_level.text().isdigit():
            return False, "سطح ولتاژ باید عدد باشد!"
        if self.operation_year.text() and not (self.operation_year.text().isdigit() and len(self.operation_year.text()) == 4):
            return False, "سال بهره‌برداری باید عدد ۴ رقمی باشد!"
        if self.bundle_count.text() and not self.bundle_count.text().isdigit():
            return False, "تعداد باندل باید عدد باشد!"
        if self.circuit_count.text() and not self.circuit_count.text().isdigit():
            return False, "تعداد مدار باید عدد باشد!"
        existing_codes = self.db.fetch_all("SELECT line_code FROM lines WHERE id != ?", (self.current_id or -1,))
        if self.line_code.text() in [code[0] for code in existing_codes]:
            return False, "کد خط باید یکتا باشد!"
        return True, ""

    def save_line(self):
        is_valid, error_msg = self.validate_inputs()
        if not is_valid:
            QMessageBox.warning(self, "خطا", error_msg)
            return

        def format_number(value):
            if value:
                try:
                    num = float(value)
                    return str(int(num)) if num.is_integer() else str(num).rstrip('0').rstrip('.')
                except ValueError:
                    return value
            return value

        self.line_data = {
            'line_code': self.line_code.text(),
            'voltage_level': self.voltage_level.text(),
            'line_name': self.line_name.text(),
            'dispatch_code': self.dispatch_code.text(),
            'total_towers': self.total_towers.text(),
            'tension_towers': self.tension_towers.text(),
            'suspension_towers': self.suspension_towers.text(),
            'line_length': format_number(self.line_length.text()),
            'circuit_length': format_number(self.circuit_length.text()),
            'circuit_count': self.circuit_count.text(),
            'plain_area': self.plain_area.text(),
            'semi_mountainous': self.semi_mountainous.text(),
            'rough_terrain': self.rough_terrain.text(),
            'supervisor': self.supervisor.text(),
            'team_leader': self.team_leader.currentText(),
            'operation_year': self.operation_year.text(),
            'wire_type': self.wire_type.text(),
            'tower_type': self.tower_type.text(),
            'bundle_count': self.bundle_count.text(),
            'id': self.current_id
        }
        self.accept()

# -----------------------------------------------
# Lines Window: Displays and manages power line information with column-specific filtering
# -----------------------------------------------

class LinesWindow(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.db = Database()
        self.init_ui()
        
    def init_ui(self):
        """Initialize the user interface"""
        self.setup_layout()
        self.setup_style()
        self.setup_widgets()
        self.setup_connections()
        self.load_table()
    
    def setup_layout(self):
        """Setup the main layout"""
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(MARGIN, MARGIN, MARGIN, MARGIN)
        self.layout.setSpacing(PADDING * 2)
        self.setLayoutDirection(Qt.RightToLeft)
        
    def setup_style(self):
        """Setup window style"""
        self.setStyleSheet("background-color: #f0f0f0;")
        
    def setup_widgets(self):
        """Setup all widgets"""
        self.setup_headers()
        self.setup_toolbar()
        self.setup_filter()
        self.setup_table()
        
    def setup_headers(self):
        """Setup header configurations"""
        # Headers setup
        self.original_headers = [
            "کد خط", "سطح ولتاژ", "نام خط", "کد دیسپاچینگ", "تعداد مدار",
            "تعداد باندل", "طول خط", "طول مدار", "تعداد کل دکل‌ها",
            "آویز", "زاویه", "دشت", "نیمه کوهستانی", "صعب العبور",
            "کارشناس خط", "نوع دکل", "نوع سیم", "سال بهره‌برداری"
        ]
        
        self.column_names = [
            "line_code", "voltage_level", "line_name", "dispatch_code", "circuit_count",
            "bundle_count", "line_length", "circuit_length", "total_towers",
            "suspension_towers", "tension_towers", "plain_area", "semi_mountainous", "rough_terrain",
            "team_leader", "tower_type", "wire_type", "operation_year"
        ]
        
        self.column_filters = {}
        
    def setup_toolbar(self):
        """Setup toolbar with actions"""
        self.toolbar = QToolBar()
        self.toolbar.setFont(QFont(FONT_FAMILY, FONT_SIZE_LARGE))
        self.toolbar.setToolButtonStyle(Qt.ToolButtonTextBesideIcon)
        self.toolbar.setStyleSheet("""
            QToolBar { 
                background-color: #e0e0e0; 
                padding: 5px; 
                spacing: 5px; 
                border-radius: 8px;
            }
            QToolButton { 
                margin: 2px; 
                background-color: transparent !important; 
                border: none; 
                padding: 2px; 
                cursor: pointer !important;
            }
            QToolButton:flat { background: transparent; }
            QToolButton:hover { 
                background-color: #d0d0d0;
                cursor: pointer !important;
                border-radius: 5px;
                padding: 5px; 
            }
            QToolSeparator { color: #333333; font-weight: bold; width: 10px; }
        """)
        
        # Create actions with standardized icons
        self.add_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_FileDialogNewFolder)), "افزودن خط جدید", self)
        self.delete_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_TrashIcon)), "حذف", self)
        self.edit_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_FileDialogContentsView)), "ویرایش", self)
        self.import_excel_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_FileDialogStart)), "ورود اطلاعات از اکسل", self)
        self.report_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_FileDialogDetailedView)), "خروجی گرفتن", self)
        self.back_action = QAction(QIcon(self.style().standardIcon(QStyle.SP_ArrowBack)), "برگشت", self)

        # Add actions to toolbar
        self.toolbar.addAction(self.add_action)
        self.toolbar.addSeparator().setText("|")
        self.toolbar.addAction(self.delete_action)
        self.toolbar.addSeparator().setText("|")
        self.toolbar.addAction(self.edit_action)
        self.toolbar.addSeparator().setText("|")
        self.toolbar.addAction(self.import_excel_action)
        self.toolbar.addSeparator().setText("|")
        self.toolbar.addAction(self.report_action)
        self.toolbar.addSeparator().setText("|")
        self.toolbar.addAction(self.back_action)
        
        self.layout.addWidget(self.toolbar)

    def setup_table(self):
        """Setup main data table"""
        self.table = QTableWidget()
        self.table.setColumnCount(18)  # Match the number of headers
        self.table.setHorizontalHeaderLabels(self.original_headers)
        self.table.setFont(QFont(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.table.setStyleSheet("""
            QTableWidget { 
                border: 1px solid #ccc; 
                background-color: white; 
            }
            QHeaderView::section {
                background-color: #f0f0f0;
                padding: 5px;
                border: 1px solid #ccc;
                font-weight: bold;
            }
        """)
        
        # Table settings
        self.table.setSelectionMode(QTableWidget.ExtendedSelection)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.setEditTriggers(QTableWidget.AnyKeyPressed | QTableWidget.SelectedClicked)
        
        # Connect signals
        self.table.cellClicked.connect(self.load_row_data)
        self.table.cellChanged.connect(self.save_cell)
        self.table.setContextMenuPolicy(Qt.CustomContextMenu)
        self.table.customContextMenuRequested.connect(self.show_context_menu)
        self.table.horizontalHeader().sectionClicked.connect(self.show_filter_popover)
        
        self.layout.addWidget(self.table)

    def setup_filter(self):
        """Setup filter section"""
        self.filter_layout = QHBoxLayout()
        self.filter_input = QLineEdit()
        self.filter_input.setPlaceholderText("جستجو در تمام ستون‌ها...")
        self.filter_input.setFont(QFont(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.filter_input.setStyleSheet("""
            QLineEdit {
                padding: 5px;
                border: 1px solid #ccc;
                border-radius: 8px;
                background-color: white;
                width: 300px;
            }
            QLineEdit:focus {
                border: 1px solid #4CAF50;
            }
        """)
        self.filter_input.setFixedWidth(300)
        
        # Connect signals
        self.filter_input.textChanged.connect(self.filter_table)
        
        # Add to layout
        filter_label = QLabel("فیلتر:")
        filter_label.setFont(QFont(FONT_FAMILY, FONT_SIZE_NORMAL))
        self.filter_layout.addWidget(filter_label)
        self.filter_layout.addWidget(self.filter_input)
        self.filter_layout.addStretch()
        
        self.layout.addLayout(self.filter_layout)

    def setup_connections(self):
        """Setup signal connections"""
        self.add_action.triggered.connect(self.add_line)
        self.delete_action.triggered.connect(self.delete_line)
        self.edit_action.triggered.connect(self.edit_line)
        self.import_excel_action.triggered.connect(self.import_from_excel)
        self.report_action.triggered.connect(self.generate_report)
        self.back_action.triggered.connect(self.close)
        
    def show_filter_popover(self, logical_index):
        try:
            header = self.table.horizontalHeader()
            viewport = self.table.viewport()
            visual_index = header.visualIndex(logical_index)
            section_pos = header.sectionViewportPosition(visual_index)
            section_width = header.sectionSize(visual_index)
            header_height = header.height()
            h_scroll = self.table.horizontalScrollBar().value()
            v_scroll = self.table.verticalScrollBar().value()
            adjusted_pos = section_pos - h_scroll
            total_width = viewport.width()
            if adjusted_pos < 0:
                adjusted_pos = 0
            elif adjusted_pos + section_width > total_width:
                adjusted_pos = total_width - section_width
            local_pos = QPoint(adjusted_pos, header_height - v_scroll)
            global_pos = self.table.mapToGlobal(local_pos)
            current_filter = self.column_filters.get(self.column_names[logical_index], "")
            popover = FilterPopover(
                self.table,
                logical_index,
                self.original_headers[logical_index],
                current_filter,
                self.update_column_filter
            )
            popover.move(global_pos)
            popover.show()
            popover.filter_input.setFocus()
            logging.debug(f"Filter popover for column {logical_index} at global_pos {global_pos}")
        except Exception as e:
            logging.error(f"Error in show_filter_popover: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در نمایش فیلتر: {str(e)}")

    def update_column_filter(self, column_index, filter_text):
        try:
            column_name = self.column_names[column_index]
            logging.debug(f"Updating filter for column {column_name}: {filter_text}")
            if filter_text:
                self.column_filters[column_name] = filter_text
            elif column_name in self.column_filters:
                del self.column_filters[column_name]
            new_headers = self.original_headers.copy()
            for i, name in enumerate(self.column_names):
                if name in self.column_filters and self.column_filters[name]:
                    new_headers[i] = f"{self.original_headers[i]} ▼"
                else:
                    new_headers[i] = self.originalHeaders[i]
            self.table.setHorizontalHeaderLabels(new_headers)
            self.load_table()
        except Exception as e:
            logging.error(f"Error in update_column_filter: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در به‌روزرسانی فیلتر: {str(e)}")

    def load_table(self, global_filter=None):
        try:
            self.table.blockSignals(True)
            query = """
                SELECT id, line_code, voltage_level, line_name, dispatch_code, circuit_count,
                       bundle_count, line_length, circuit_length, total_towers,
                       suspension_towers, tension_towers, plain_area, semi_mountainous,
                       rough_terrain, team_leader, tower_type, wire_type, operation_year
                FROM lines
            """
            params = []
            conditions = []
            if global_filter:
                conditions.append("""
                    (line_code LIKE ? OR voltage_level LIKE ? OR line_name LIKE ? 
                    OR dispatch_code LIKE ? OR circuit_count LIKE ? OR bundle_count LIKE ?
                    OR line_length LIKE ? OR circuit_length LIKE ? OR total_towers LIKE ?
                    OR suspension_towers LIKE ? OR tension_towers LIKE ? OR plain_area LIKE ?
                    OR semi_mountainous LIKE ? OR rough_terrain LIKE ? OR team_leader LIKE ?
                    OR tower_type LIKE ? OR wire_type LIKE ? OR operation_year LIKE ?)
                """)
                params.extend([f"%{global_filter}%"] * 18)
            for column, filter_text in self.column_filters.items():
                if filter_text:
                    conditions.append(f"{column} LIKE ?")
                    params.append(f"%{filter_text}%")
            if conditions:
                query += " WHERE " + " AND ".join(conditions)
            logging.debug(f"Executing query: {query} with params: {params}")
            rows = self.db.fetch_all(query, tuple(params))
            self.table.setRowCount(len(rows))
            for row_idx, row_data in enumerate(rows):
                for col_idx, data in enumerate(row_data[1:], start=0):
                    data = str(data) if data is not None else ""
                    if col_idx in [7, 8]:
                        try:
                            num = float(data)
                            data = str(int(num)) if num.is_integer() else str(num).rstrip('0').rstrip('.')
                        except (ValueError, TypeError):
                            pass
                    item = QTableWidgetItem(data)
                    item.setTextAlignment(Qt.AlignCenter)
                    self.table.setItem(row_idx, col_idx, item)
                self.table.item(row_idx, 0).setData(Qt.UserRole, row_data[0])
            self.table.resizeColumnsToContents()
            for col in range(self.table.columnCount()):
                if self.table.columnWidth(col) < 100:
                    self.table.setColumnWidth(col, 100)
            self.table.resizeRowsToContents()
            self.table.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            logging.debug(f"Table loaded with {len(rows)} rows")
        except Exception as e:
            logging.error(f"Error in load_table: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در بارگذاری جدول: {str(e)}")
        finally:
            self.table.blockSignals(False)

    def filter_table(self):
        try:
            self.load_table(self.filter_input.text())
        except Exception as e:
            logging.error(f"Error in filter_table: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در فیلتر جدول: {str(e)}")

    def add_line(self):
        dialog = LineInputDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            line_data = dialog.line_data
            try:
                self.db.execute_query(
                    """
                    INSERT INTO lines (line_code, voltage_level, line_name, dispatch_code, circuit_count,
                                       bundle_count, line_length, circuit_length, total_towers,
                                       suspension_towers, tension_towers, plain_area, semi_mountainous, rough_terrain,
                                       supervisor, team_leader, tower_type, wire_type, operation_year)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        line_data['line_code'], line_data['voltage_level'], line_data['line_name'], 
                        line_data['dispatch_code'], line_data['circuit_count'],
                        line_data['bundle_count'], line_data['line_length'], line_data['circuit_length'],
                        line_data['total_towers'],
                        line_data['suspension_towers'], line_data['tension_towers'], line_data['plain_area'],
                        line_data['semi_mountainous'], line_data['rough_terrain'], line_data['supervisor'],
                        line_data['team_leader'], line_data['tower_type'], line_data['wire_type'],
                        line_data['operation_year']
                    )
                )
                self.load_table()
                QMessageBox.information(self, "موفقیت", "خط با موفقیت اضافه شد.")
            except Exception as e:
                logging.error(f"Error in add_line: {str(e)}\n{traceback.format_exc()}")
                QMessageBox.critical(self, "خطا", f"خطا در افزودن خط: {str(e)}")

    def edit_line(self):
        try:
            selected_items = self.table.selectedItems()
            if not selected_items:
                QMessageBox.warning(self, "هشدار", "لطفاً یک ردیف را برای ویرایش انتخاب کنید.")
                return

            selected_row = selected_items[0].row()
            line_code = self.table.item(selected_row, self.column_names.index("line_code")).text()

            # Get data from input fields
            line_data = {
                "voltage_level": self.voltage_level_input.text(),
                "line_name": self.line_name_input.text(),
                "dispatch_code": self.dispatch_code_input.text(),
                "circuit_count": self.circuit_count_input.text(),
                "bundle_count": self.bundle_count_input.text(),
                "line_length": self.line_length_input.text(),
                "circuit_length": self.circuit_length_input.text(),
                "total_towers": self.total_towers_input.text(),
                "suspension_towers": self.suspension_towers_input.text(),
                "tension_towers": self.tension_towers_input.text(),
                "plain_area": self.plain_area_input.text(),
                "semi_mountainous": self.semi_mountainous_input.text(),
                "rough_terrain": self.rough_terrain_input.text(),
                "supervisor": self.supervisor_input.text(),
                "team_leader": self.team_leader_input.text(),
                "tower_type": self.tower_type_input.text(),
                "wire_type": self.wire_type_input.text(),
                "operation_year": self.operation_year_input.text()
            }

            # Create SQL query
            update_pairs = [f"{key} = ?" for key in line_data.keys()]
            query = f"UPDATE lines SET {', '.join(update_pairs)} WHERE line_code = ?"

            # Execute query with data
            values = list(line_data.values())
            values.append(line_code)  # Add line_code for WHERE clause
            self.cursor.execute(query, tuple(values))
            self.conn.commit()
            self.load_table()

            # Clear input fields
            self.clear_inputs()

            QMessageBox.information(self, "موفقیت", "اطلاعات خط با موفقیت به‌روزرسانی شد.")

        except Exception as e:
            QMessageBox.critical(self, "خطا", f"خطا در به‌روزرسانی اطلاعات: {str(e)}")
            print(f"Error in edit_line: {str(e)}")

    def delete_line(self):
        selected_rows = sorted(set(index.row() for index in self.table.selectedIndexes()))
        if not selected_rows:
            QMessageBox.warning(self, "خطا", "خطی انتخاب نشده است!")
            return
        row_count = len(selected_rows)
        msg = f"آیا از حذف {row_count} خط مطمئن هستید؟"
        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("تأیید حذف")
        msg_box.setText(msg)
        msg_box.setStandardButtons(QMessageBox.NoButton)
        yes_button = msg_box.addButton("بله", QMessageBox.YesRole)
        no_button = msg_box.addButton("خیر", QMessageBox.RejectRole)
        yes_button.setFont(QFont("Vazir", 12))
        no_button.setFont(QFont("Vazir", 12))
        msg_box.setDefaultButton(no_button)
        msg_box.setLayoutDirection(Qt.RightToLeft)
        msg_box.exec_()
        if msg_box.clickedButton() != yes_button:
            return
        try:
            for row in selected_rows:
                line_code = self.table.item(row, 0).text() if self.table.item(row, 0) else ""
                self.db.execute_query("DELETE FROM lines WHERE line_code=?", (line_code,))
            self.load_table()
            QMessageBox.information(self, "موفقیت", f"{row_count} خط با موفقیت حذف شد.")
        except Exception as e:
            logging.error(f"Error in delete_line: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در حذف: {str(e)}")

    def load_row_data(self):
        pass
        
    def generate_report(self):
        """Generate Excel report with proper formatting"""
        try:
            query = """
                SELECT 
                    line_code, voltage_level, line_name, dispatch_code, circuit_count,
                    bundle_count, line_length, circuit_length, total_towers,
                    suspension_towers, tension_towers, plain_area, semi_mountainous, 
                    rough_terrain, team_leader, tower_type, wire_type, operation_year
                FROM lines
                ORDER BY line_code
            """
            rows = self.db.fetch_all(query)

            if not rows:
                QMessageBox.warning(self, "هشدار", "هیچ داده‌ای برای گزارش‌گیری وجود ندارد.")
                return

            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'گزارش خطوط'
            
            # Write headers
            for col, header in enumerate(self.original_headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(name=FONT_FAMILY, bold=True, size=12)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Write data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value is not None else "")
                    cell.font = openpyxl.styles.Font(name=FONT_FAMILY, size=11)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths (multiply by 2 for Persian characters)
            for col in ws.columns:
                max_length = 0
                column = get_column_letter(col[0].column)
                
                for cell in col:
                    try:
                        max_length = max(max_length, len(str(cell.value)) * 2)
                    except:
                        pass
                
                ws.column_dimensions[column].width = max_length
            
            # Set RTL direction
            ws.sheet_view.rightToLeft = True
            
            # Get save location
            file_path, _ = QFileDialog.getSaveFileName(
                self,
                "ذخیره فایل گزارش",
                os.path.join("Data", "گزارش خطوط.xlsx"),
                "Excel Files (*.xlsx)"
            )
            
            if file_path:
                wb.save(file_path)
                QMessageBox.information(self, "موفقیت", "گزارش با موفقیت ایجاد شد.")

        except Exception as e:
            logging.error(f"Error in generate_report: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در ایجاد گزارش: {str(e)}")

    def format_number(self, value):
        if pd.notna(value):
            try:
                num = float(value)
                return str(int(num)) if num.is_integer() else str(num).rstrip('0').rstrip('.')
            except (ValueError, TypeError):
                return str(value)
        return ""

    def import_from_excel(self):
        try:
            file_path, _ = QFileDialog.getOpenFileName(self, "انتخاب فایل اکسل", "", "Excel Files (*.xlsx *.xls)")
            if not file_path:
                return

            df = pd.read_excel(file_path)
            logging.debug(f"Columns found in Excel file: {list(df.columns)}")

            excel_columns = {
                "کد خط": "line_code",
                "سطح ولتاژ": "voltage_level",
                "نام خط": "line_name",
                "کد دیسپاچینگ": "dispatch_code",
                "تعداد مدار": "circuit_count",
                "تعداد باندل": "bundle_count",
                "طول خط": "line_length",
                "طول مدار": "circuit_length",
                "تعداد کل دکل‌ها": "total_towers",
                "آویز": "suspension_towers",
                "زاویه": "tension_towers",
                "دشت": "plain_area",
                "نیمه کوهستانی": "semi_mountainous",
                "صعب العبور": "rough_terrain",
                "کارشناس خط": "team_leader",
                "سرپرست خط": "tower_type",
                "نوع دکل": "wire_type",
                "سال بهره‌برداری": "operation_year"
            }

            missing_columns = [col for col in excel_columns if col not in df.columns]
            if missing_columns:
                error_msg = "ستون‌های ضروری زیر در فایل اکسل یافت نشدند:\n"
                for col in missing_columns:
                    error_msg += f"- {col}\n"
                QMessageBox.critical(self, "خطا", error_msg)
                return

            existing_codes = {row[0] for row in self.db.fetch_all("SELECT line_code FROM lines") if row[0]}
            inserted_count = 0
            errors = []

            for index, row in df.iterrows():
                data = {}
                for excel_col, db_col in excel_columns.items():
                    data[db_col] = str(row[excel_col]) if pd.notna(row[excel_col]) else ""

                if not data["line_code"]:
                    errors.append(f"ردیف {index + 2}: کد خط خالی است")
                    continue
                if data["line_code"] in existing_codes:
                    errors.append(f"ردیف {index + 2}: کد خط '{data['line_code']}' تکراری است")
                    continue
                if not data["line_name"]:
                    errors.append(f"ردیف {index + 2}: نام خط خالی است")
                    continue

                if data["total_towers"] and not data["total_towers"].isdigit():
                    errors.append(f"ردیف {index + 2}: تعداد کل دکل‌ها باید عدد باشد")
                    continue

                # Insert valid rows into the database
                try:
                    self.db.execute_query(
                        """
                        INSERT INTO lines (line_code, voltage_level, line_name, dispatch_code, circuit_count,
                                           bundle_count, line_length, circuit_length, total_towers,
                                           suspension_towers, tension_towers, plain_area, semi_mountainous, rough_terrain,
                                           team_leader, tower_type, wire_type, operation_year)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            data["line_code"], data["voltage_level"], data["line_name"], data["dispatch_code"],
                            data["circuit_count"], data["bundle_count"], data["line_length"], data["circuit_length"],
                            data["total_towers"], data["suspension_towers"], data["tension_towers"], data["plain_area"],
                            data["semi_mountainous"], data["rough_terrain"], data["team_leader"], data["tower_type"],
                            data["wire_type"], data["operation_year"]
                        )
                    )
                    inserted_count += 1
                except Exception as e:
                    errors.append(f"ردیف {index + 2}: خطا در درج اطلاعات: {str(e)}")

            # Show results
            if errors:
                error_msg = "خطاهای زیر در ورود اطلاعات رخ دادند:\n"
                for error in errors:
                    error_msg += f"- {error}\n"
                QMessageBox.warning(self, "خطاها", error_msg)

            QMessageBox.information(self, "نتیجه", f"{inserted_count} ردیف با موفقیت وارد شدند.")

            self.load_table()

        except Exception as e:
            logging.error(f"Error in import_from_excel: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در ورود اطلاعات از اکسل: {str(e)}")
        self.load_table()

    def show_context_menu(self, position):
        try:
            menu = QMenu(self)
            menu.setStyleSheet("""
                QMenu { 
                    background-color: white; 
                    border: 1px solid #ccc; 
                    direction: ltr; 
                }
                QMenu::item { 
                    padding: 5px 20px; 
                    color: black; 
                    direction: ltr; 
                }
                QMenu::item:selected { 
                    background-color: #d0d0d0; 
                    color: black; 
                }
            """)
            copy_action = QAction("کپی", self)
            copy_action.triggered.connect(self.copy_selected_cells)
            menu.addAction(copy_action)
            edit_action = QAction("ویرایش", self)
            edit_action.triggered.connect(self.edit_line)
            menu.addAction(edit_action)
            delete_action = QAction("حذف", self)
            delete_action.triggered.connect(self.delete_line)
            menu.addAction(delete_action)
            selected_items = bool(self.table.selectedItems())
            copy_action.setEnabled(selected_items)
            edit_action.setEnabled(selected_items)
            delete_action.setEnabled(selected_items)
            menu.exec_(self.table.viewport().mapToGlobal(position))
        except Exception as e:
            logging.error(f"Error in show_context_menu: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در نمایش منوی راست‌کلیک: {str(e)}")

    def copy_selected_cells(self):
        selected_items = self.table.selectedItems()
        if not selected_items:
            return
        rows = sorted(set(item.row() for item in selected_items))
        cols = sorted(set(item.column() for item in selected_items))
        text = ""
        for row in rows:
            row_data = []
            for col in cols:
                item = self.table.item(row, col)
                row_data.append(item.text() if item else "")
            text += "\t".join(row_data) + "\n"
        clipboard = QApplication.clipboard()
        clipboard.setText(text)
        QMessageBox.information(self, "موفقیت", "محتوای سلول‌های انتخاب‌شده کپی شد.")

    def save_cell(self, row, col):
        try:
            item = self.table.item(row, col)
            if not item:
                return
            new_value = item.text().strip()
            line_id = self.table.item(row, 0).data(Qt.UserRole)
            if not line_id:
                logging.error("No line_id found for row")
                return
            column_name = self.column_names[col]
            if column_name in ['total_towers', 'tension_towers', 'suspension_towers', 'bundle_count', 'circuit_count']:
                if new_value and not new_value.isdigit():
                    QMessageBox.critical(self, "خطا", f"مقدار {self.original_headers[col]} باید عدد باشد!")
                    self.load_table()
                    return
            elif column_name in ['line_length', 'circuit_length']:
                if new_value and not re.match(r"^\d*\.?\d*$", new_value):
                    QMessageBox.critical(self, "خطا", f"مقدار {self.original_headers[col]} باید عدد باشد!")
                    self.load_table()
                    return
            elif column_name == 'voltage_level':
                if new_value and not new_value.isdigit():
                    QMessageBox.critical(self, "خطا", "سطح ولتاژ باید عدد باشد!")
                    self.load_table()
                    return
            elif column_name == 'operation_year':
                if new_value and not (new_value.isdigit() and len(new_value) == 4):
                    QMessageBox.critical(self, "خطا", "سال باید عدد ۴ رقمی باشد!")
                    self.load_table()
                    return
            elif column_name == 'line_code':
                if not new_value:
                    QMessageBox.critical(self, "خطا", "کد خط اجباری است!")
                    self.load_table()
                    return
                existing_codes = self.db.fetch_all("SELECT line_code FROM lines WHERE id != ?", (line_id,))
                if new_value in [code[0] for code in existing_codes]:
                    QMessageBox.critical(self, "خطا", "کد خط باید یکتا باشد!")
                    self.load_table()
                    return
            elif column_name == 'line_name':
                if not new_value:
                    QMessageBox.critical(self, "خطا", "نام خط اجباری است!")
                    self.load_table()
                    return
            elif column_name == 'team_leader':
                if not new_value:
                    QMessageBox.critical(self, "خطا", "سرپرست اکیپ اجباری است!")
                    self.load_table()
                    return
                valid_leaders = self.db.fetch_all("SELECT first_name || ' ' || last_name FROM teams WHERE position = 'سرپرست اکیپ'")
                if new_value not in [row[0] for row in valid_leaders]:
                    QMessageBox.critical(self, "خطا", "سرپرست اکیپ نامعتبر است!")
                    self.load_table()
                    return
            query = f"UPDATE lines SET {column_name} = ? WHERE id = ?"
            self.db.execute_query(query, (new_value, line_id))
            logging.debug(f"Updated {column_name} to '{new_value}' for line_id {line_id}")
        except Exception as e:
            logging.error(f"Error in save_cell: {str(e)}\n{traceback.format_exc()}")
            QMessageBox.critical(self, "خطا", f"خطا در ذخیره تغییرات: {str(e)}")
            self.load_table()

    def clear_inputs(self):
        self.line_code_input.clear()
        self.voltage_level_input.clear()
        self.line_name_input.clear()
        self.dispatch_code_input.clear()
        self.circuit_count_input.clear()
        self.bundle_count_input.clear()
        self.line_length_input.clear()
        self.circuit_length_input.clear()
        self.total_towers_input.clear()
        self.suspension_towers_input.clear()
        self.tension_towers_input.clear()
        self.plain_area_input.clear()
        self.semi_mountainous_input.clear()
        self.rough_terrain_input.clear()
        self.supervisor_input.clear()
        self.team_leader_input.clear()
        self.tower_type_input.clear()
        self.wire_type_input.clear()
        self.operation_year_input.clear()

    # -----------------------------------------------
    # Line Input Dialog: A dialog for adding new lines with validation and standardized styling
    # -----------------------------------------------

    class LineInputDialog(QDialog):
        def __init__(self, parent=None):
            super().__init__(parent)
            self.setWindowTitle("افزودن خط جدید")
            self.setup_ui()
            
        def setup_ui(self):
            """Initialize dialog UI"""
            self.layout = QVBoxLayout(self)
            self.layout.setContentsMargins(MARGIN, MARGIN, MARGIN, MARGIN)
            self.layout.setSpacing(PADDING)
            
            # Form layout for input fields
            self.form_layout = QFormLayout()
            self.form_layout.setSpacing(PADDING)
            
            # Create input fields
            self.create_input_fields()
            
            # Add form layout to main layout
            self.layout.addLayout(self.form_layout)
            
            # Add spacing
            self.layout.addSpacerItem(QSpacerItem(0, 15, QSizePolicy.Minimum, QSizePolicy.Fixed))
            
            # Buttons
            self.button_layout = QHBoxLayout()
            self.save_button = QPushButton("ذخیره")
            self.cancel_button = QPushButton("لغو")
            
            self.save_button.setStyleSheet("""
                QPushButton {
                    background-color: #4CAF50;
                    color: white;
                    border: none;
                    padding: 5px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #45a049;
                }
            """)
            
            self.cancel_button.setStyleSheet("""
                QPushButton {
                    background-color: #f44336;
                    color: white;
                    border: none;
                    padding: 5px 15px;
                    border-radius: 4px;
                }
                QPushButton:hover {
                    background-color: #da190b;
                }
            """)
            
            self.button_layout.addWidget(self.save_button)
            self.button_layout.addWidget(self.cancel_button)
            self.layout.addLayout(self.button_layout)
            
            # Connect signals
            self.save_button.clicked.connect(self.accept)
            self.cancel_button.clicked.connect(self.reject)
            
        def create_input_fields(self):
            """Create and setup input fields"""
            # Create input fields with proper validation
            self.line_code = QLineEdit()
            self.voltage_level = QLineEdit()
            self.line_name = QLineEdit()
            self.dispatch_code = QLineEdit()
            self.circuit_count = QLineEdit()
            self.bundle_count = QLineEdit()
            self.line_length = QLineEdit()
            self.circuit_length = QLineEdit()
            self.total_towers = QLineEdit()
            self.suspension_towers = QLineEdit()
            self.tension_towers = QLineEdit()
            self.plain_area = QLineEdit()
            self.semi_mountainous = QLineEdit()
            self.rough_terrain = QLineEdit()
            self.team_leader = QLineEdit()
            self.tower_type = QLineEdit()
            self.wire_type = QLineEdit()
            self.operation_year = QLineEdit()
            
            # Set font for all fields
            for field in self.__dict__.values():
                if isinstance(field, QLineEdit):
                    field.setFont(QFont(FONT_FAMILY, FONT_SIZE_NORMAL))
            
            # Add fields to form layout with labels
            field_labels = [
                ("کد خط:", self.line_code),
                ("سطح ولتاژ:", self.voltage_level),
                ("نام خط:", self.line_name),
                ("کد دیسپاچینگ:", self.dispatch_code),
                ("تعداد مدار:", self.circuit_count),
                ("تعداد باندل:", self.bundle_count),
                ("طول خط:", self.line_length),
                ("طول مدار:", self.circuit_length),
                ("تعداد کل دکل‌ها:", self.total_towers),
                ("آویز:", self.suspension_towers),
                ("زاویه:", self.tension_towers),
                ("دشت:", self.plain_area),
                ("نیمه کوهستانی:", self.semi_mountainous),
                ("صعب العبور:", self.rough_terrain),
                ("کارشناس خط:", self.team_leader),
                ("نوع دکل:", self.tower_type),
                ("نوع سیم:", self.wire_type),
                ("سال بهره‌برداری:", self.operation_year)
            ]
            
            for label, field in field_labels:
                label_widget = QLabel(label)
                label_widget.setFont(QFont(FONT_FAMILY, FONT_SIZE_NORMAL))
                self.form_layout.addRow(label_widget, field)
                
        def validate_input(self):
            """Validate input fields"""
            if not self.line_code.text():
                QMessageBox.warning(self, "خطا", "کد خط نمی‌تواند خالی باشد.")
                return False
                
            if not self.line_name.text():
                QMessageBox.warning(self, "خطا", "نام خط نمی‌تواند خالی باشد.")
                return False
                
            if self.line_length.text() and not re.match(r"^\d*\.?\d*$", self.line_length.text()):
                QMessageBox.warning(self, "خطا", "طول خط باید عدد باشد.")
                return False
                
            if self.circuit_length.text() and not re.match(r"^\d*\.?\d*$", self.circuit_length.text()):
                QMessageBox.warning(self, "خطا", "طول مدار باید عدد باشد.")
                return False
                
            return True
            
        def get_line_data(self):
            """Get the line data as a dictionary"""
            if not self.validate_input():
                return None
                
            return {
                'line_code': self.line_code.text(),
                'voltage_level': self.voltage_level.text(),
                'line_name': self.line_name.text(),
                'dispatch_code': self.dispatch_code.text(),
                'circuit_count': self.circuit_count.text(),
                'bundle_count': self.bundle_count.text(),
                'line_length': self.line_length.text(),
                'circuit_length': self.circuit_length.text(),
                'total_towers': self.total_towers.text(),
                'suspension_towers': self.suspension_towers.text(),
                'tension_towers': self.tension_towers.text(),
                'plain_area': self.plain_area.text(),
                'semi_mountainous': self.semi_mountainous.text(),
                'rough_terrain': self.rough_terrain.text(),
                'team_leader': self.team_leader.text(),
                'tower_type': self.tower_type.text(),
                'wire_type': self.wire_type.text(),
                'operation_year': self.operation_year.text()
            }

    def clear_placeholder(self, event):
        """Clear placeholder text when the filter input is clicked"""
        try:
            if self.filter_input.placeholderText() and not self.filter_input.text():
                self.filter_input.setPlaceholderText("")
            super(QLineEdit, self.filter_input).mousePressEvent(event)
        except Exception as e:
            logging.error(f"Error in clear_placeholder: {str(e)}\n{traceback.format_exc()}")
